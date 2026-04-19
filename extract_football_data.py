import openpyxl
import json

# Load the workbook
workbook = openpyxl.load_workbook('football_analysis.xlsx')

# Dictionary to hold the data
data = {}

# Extracting sheet names, dimensions, headers, and data
for sheet in workbook.sheetnames:
    ws = workbook[sheet]
    dimensions = ws.dimensions
    headers = [cell.value for cell in ws[1]]  # Assuming the first row contains headers
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
        rows.append(row)
    data[sheet] = {
        'dimensions': dimensions,
        'headers': headers,
        'data': rows
    }

# Save the extracted data to a JSON file
with open('football_analysis.json', 'w') as json_file:
    json.dump(data, json_file, indent=4)